import math
import openpyxl


class RealData:
    def __init__(self):
        self._freq = 0
        self._z_re = 0
        self._z_img = 0
        self._s_11_re = 0
        self._s_11_img = 0
        self._s_11_db = 0
        self._s_12_db = 0

    def set_freq(self, f: float) -> None:
        self._freq = f

    def set_z_re(self, zre: float) -> None:
        self._z_re = zre

    def set_z_img(self, zimg: float) -> None:
        self._z_img = zimg

    def set_s_11_db(self, s11db: float) -> None:
        self._s_11_db = s11db

    def set_s_11_re(self, s11re: float) -> None:
        self._s_11_re = s11re

    def set_s_11_img(self, s11img: float) -> None:
        self._s_11_img = s11img

    def set_s_12_db(self, s12db: float) -> None:
        self._s_12_db = s12db

    def get_freq(self) -> float:
        return self._freq

    def get_z_re(self) -> float:
        return self._z_re

    def get_z_img(self) -> float:
        return self._z_img

    def get_s_11_db(self) -> float:
        return self._s_11_db

    def get_s_12_db(self) -> float:
        return self._s_12_db

    def calc_s_11_db(self) -> None:
        self._s_11_db = 20 + math.log10(math.sqrt(self._s_11_re**2 + self._s_11_img**2))


class FarfieldDirectivity:
    def __init__(self):
        self._angle = 0
        self._h_s_21_db = 0
        self._h_s_21_norm = 0
        self._v_s_21_db = 0
        self._v_s_21_norm = 0
        self._v_dir = 0
        self._h_dir = 0

    def set_angle(self, a: float) -> None:
        self._angle = a

    def set_h_s_21_db(self, hs21db: float) -> None:
        self._h_s_21_db = hs21db

    def set_h_s_21_norm(self, hs21norm: float) -> None:
        self._h_s_21_norm = hs21norm

    def set_v_s_21_db(self, vs21db: float) -> None:
        self._v_s_21_db = vs21db

    def set_v_s_21_norm(self, vs21norm: float) -> None:
        self._v_s_21_norm = vs21norm

    def set_v_dir(self, vdir: float) -> None:
        self._v_dir = vdir

    def set_h_dir(self, hdir: float) -> None:
        self._h_dir = hdir

    def get_angle(self) -> float:
        return self._angle

    def get_h_s_21_db(self) -> float:
        return self._h_s_21_db

    def get_h_s_21_norm(self) -> float:
        return self._h_s_21_norm

    def get_v_s_21_db(self) -> float:
        return self._v_s_21_db

    def get_v_s_21_norm(self) -> float:
        return self._v_s_21_norm

    def get_v_dir(self) -> float:
        return self._v_dir

    def get_h_dir(self) -> float:
        return self._h_dir


def main():
    first_line = True
    freq_results = []
    angle_results = []
    with open("InputTextFiles/MeasuredData/data.txt", 'r') as file:
        for line in file:
            tmp = line.strip().split("\t")
            if len(tmp) == 11:
                if not first_line:
                    print(tmp, len(tmp))
                    meas = RealData()
                    meas.set_freq(float(tmp[1]))
                    meas.set_s_11_db(float(tmp[2]))
                    meas.set_z_re(float(tmp[4]))
                    meas.set_z_img(float(tmp[5]))
                    meas.set_s_12_db(float(tmp[10]))
                    freq_results.append(meas)
                else:
                    first_line = False
    file.close()

    with open("InputTextFiles/MeasuredData/Phi90_vertical.txt") as file:
        for line in file:
            tmp = line.strip().split("\t")
            if len(tmp) == 3:
                myclass = FarfieldDirectivity()
                myclass.set_angle(float(tmp[0]))
                myclass.set_v_dir(float(tmp[2]))
                angle_results.append(myclass)
    file.close()

    cnt = 0
    with open("InputTextFiles/MeasuredData/Theta90_horizzontal.txt") as file:
        for line in file:
            tmp = line.strip().split("\t")
            if len(tmp) == 3:
                if angle_results[cnt].get_angle() == float(tmp[0]):
                    angle_results[cnt].set_h_dir(float(tmp[2]))
                    cnt += 1
    file.close()

    cnt = 0
    with open("InputTextFiles/MeasuredData/horizzontal.txt") as file:
        for line in file:
            tmp = line.strip().split("\t")
            print(tmp, len(tmp), cnt, angle_results[2*(cnt - 1)].get_angle())
            if len(tmp) == 5:
                if cnt == 0:
                    pass
                elif cnt == 181:
                    angle_results[-1].set_h_s_21_db(float(tmp[2]))
                    angle_results[-1].set_h_s_21_norm(float(tmp[4]))
                else:
                    if angle_results[2*(cnt - 1)].get_angle() == float(tmp[0]):
                        angle_results[2*(cnt - 1)].set_h_s_21_db(float(tmp[2]))
                        angle_results[2*(cnt - 1)].set_h_s_21_norm(float(tmp[4]))
                cnt += 1
    file.close()

    cnt = 0
    with open("InputTextFiles/MeasuredData/vertical.txt") as file:
        for line in file:
            tmp = line.strip().split("\t")
            print(tmp, len(tmp), cnt, angle_results[2*(cnt - 1)].get_angle())
            if len(tmp) == 5:
                if cnt == 0:
                    pass
                elif cnt == 181:
                    angle_results[-1].set_v_s_21_db(float(tmp[2]))
                    angle_results[-1].set_v_s_21_norm(float(tmp[4]))
                else:
                    if angle_results[2*(cnt - 1)].get_angle() == float(tmp[0]):
                        angle_results[2*(cnt - 1)].set_v_s_21_db(float(tmp[2]))
                        angle_results[2*(cnt - 1)].set_v_s_21_norm(float(tmp[4]))
                cnt += 1
    file.close()

    for el in freq_results:
        pass
        # print(el.get_freq(), el.get_s_11_db(), el.get_z_re(), el.get_z_img())

    for i in range(0, 360, 2):
        val = angle_results[i].get_h_s_21_db() + 0.5 * (angle_results[i+2].get_h_s_21_db() - angle_results[i].get_h_s_21_db())
        angle_results[i+1].set_h_s_21_db(val)
        val = angle_results[i].get_h_s_21_norm() + 0.5 * (angle_results[i+2].get_h_s_21_norm() - angle_results[i].get_h_s_21_norm())
        angle_results[i+1].set_h_s_21_norm(val)
        val = angle_results[i].get_v_s_21_db() + 0.5 * (angle_results[i+2].get_v_s_21_db() - angle_results[i].get_v_s_21_db())
        angle_results[i+1].set_v_s_21_db(val)
        val = angle_results[i].get_v_s_21_norm() + 0.5 * (angle_results[i+2].get_v_s_21_norm() - angle_results[i].get_v_s_21_norm())
        angle_results[i+1].set_v_s_21_norm(val)


    for el in angle_results:
        # pass
        print(el.get_angle(), el.get_h_dir(), el.get_v_dir(), el.get_h_s_21_db(), el.get_v_s_21_db())

    wb = openpyxl.Workbook()
    wb.create_sheet("frequencies", 0)
    wb.create_sheet("angles", 1)
    wfreq = wb["frequencies"]
    wangle = wb["angles"]

    wfreq.cell(row=1, column=1, value="Frequency [GHz]")
    wfreq.cell(row=1, column=2, value="Z Real [Ohms]")
    wfreq.cell(row=1, column=3, value="Z Imaginary [Ohms]")
    wfreq.cell(row=1, column=4, value="S11 [dB]")
    wfreq.cell(row=1, column=5, value="S21 [dB]")

    counter = 2

    for el in freq_results:
        wfreq.cell(row=counter, column=1, value=el.get_freq())
        wfreq.cell(row=counter, column=2, value=el.get_z_re())
        wfreq.cell(row=counter, column=3, value=el.get_z_img())
        wfreq.cell(row=counter, column=4, value=el.get_s_11_db())
        wfreq.cell(row=counter, column=5, value=el.get_s_12_db())
        counter += 1

    wangle.cell(row=1, column=1, value="Angle [deg]")
    wangle.cell(row=1, column=2, value="Horizzontal S21 [dB]")
    wangle.cell(row=1, column=3, value="Horizzontal S21 Normalized [dB]")
    wangle.cell(row=1, column=4, value="Horizzontal Directivity [dB]")
    wangle.cell(row=1, column=5, value="Vertical S21 [dB]")
    wangle.cell(row=1, column=6, value="Vertical S21 Normalized [dB]")
    wangle.cell(row=1, column=7, value="Vertical Directivity [dB]")

    counter = 2

    for el in angle_results:
        wangle.cell(row=counter, column=1, value=el.get_angle())
        wangle.cell(row=counter, column=2, value=el.get_h_s_21_db())
        wangle.cell(row=counter, column=3, value=el.get_h_s_21_norm())
        wangle.cell(row=counter, column=4, value=el.get_h_dir())
        wangle.cell(row=counter, column=5, value=el.get_v_s_21_db())
        wangle.cell(row=counter, column=6, value=el.get_v_s_21_norm())
        wangle.cell(row=counter, column=7, value=el.get_v_dir())
        counter += 1

    wb.save("Real Data.xlsx")


if __name__ == "__main__":
    main()