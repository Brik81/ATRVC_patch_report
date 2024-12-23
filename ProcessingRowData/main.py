# image zero -> S, VSVR and Z for an antenna with no slots

import math
import openpyxl

results = []

# S_PAR_PATH = "InputTextFiles/WithSlots/s_par_L_slot_7.6.txt"
# S_PAR_PATH = "InputTextFiles/BeforeSlots/s_par_after_deembed_corrected.txt"
S_PAR_PATH = "InputTextFiles/FirstAttemptSlots/s_par_slot_9_2.txt"

# Z_PATH = "InputTextFiles/WithSlots/Z_par_L_slot_7.6_minimum.txt"
# Z_PATH = "InputTextFiles/BeforeSlots/z_par_after_deembed_corrected.txt"
Z_PATH = "InputTextFiles/FirstAttemptSlots/z_par_slot_9_2.txt"

# VSWR_PATH = "InputTextFiles/WithSlots/VSWR_par_L_slot_7_6.txt"
# VSWR_PATH = "InputTextFiles/BeforeSlots/vswr_par_after_deembed_corrected.txt"
VSWR_PATH = "InputTextFiles/BeforeSlots/vswr_par_after_deembed_corrected.txt"

# OUT_PATH = "values_before_slots.xlsx"
# OUT_PATH = "values_with_slots.xlsx"
OUT_PATH = "values_first_attempt.xlsx"


class AntennaData:
    def __init__(self) -> None:
        self._frequency = 0
        self._s_par_re = 0
        self._s_par_img = 0
        self._s_par_mag = 0
        self._z_re = 0
        self._z_img = 0
        self._z_mag = 0
        self._vsvr = 0

    def set_freq(self, freq: float) -> None:
        self._frequency = freq

    def set_s_par_re(self, re: float) -> None:
        self._s_par_re = re

    def set_s_par_img(self, img: float) -> None:
        self._s_par_img = img

    def set_z_re(self, re: float) -> None:
        self._z_re = re

    def set_z_img(self, img: float) -> None:
        self._z_img = img

    def set_vsvr(self, vsvr: float) -> None:
        self._vsvr = vsvr

    def calc_s_par_magnitude(self) -> None:
        self._s_par_mag = 20 * math.log10(math.sqrt(self._s_par_re**2 + self._s_par_img**2))

    def calc_z_magnitude(self) -> None:
        self._z_mag = 20 * math.log10(math.sqrt(self._z_re**2 + self._z_img**2))

    def get_freq(self) -> float:
        return self._frequency

    def get_s_par_re(self) -> float:
        return self._s_par_re

    def get_s_par_img(self) -> float:
        return self._s_par_img

    def get_s_par_magnitude(self) -> float:
        return self._s_par_mag

    def get_z_re(self) -> float:
        return self._z_re

    def get_z_img(self) -> float:
        return self._z_img

    def get_z_magnitude(self) -> float:
        return self._z_mag

    def get_vsvr(self) -> float:
        return self._vsvr


def main():
    # Processing first folder
    # S Parameter
    with open(S_PAR_PATH, 'r') as file:
        for line in file:
            # line is valid if .strip("\t") returns something with len = 5
            tmp = line.strip().split("\t")
            if len(tmp) == 5:
                tmp_class = AntennaData()
                # frequency assignment
                tmp_class.set_freq(float(tmp[0]))
                # real part assignment
                tmp_class.set_s_par_re(float(tmp[1]))
                # imaginary part assignment
                tmp_class.set_s_par_img(float(tmp[2]))
                # magnitude in dB
                tmp_class.calc_s_par_magnitude()

                # add class into list
                results.append(tmp_class)
                print("Freq: %f \t Magnitute: %f\n" % (tmp_class.get_freq(), tmp_class.get_s_par_magnitude()))

        file.close()

        # find max s value and print s magnitude and frequency
        max_s_par_mag_value = float("+inf")
        class_re = 0
        for i in results:
            if i.get_s_par_magnitude() < max_s_par_mag_value:
                class_re = i
                max_s_par_mag_value = i.get_s_par_magnitude()
        print("Max value of S par is at Freq: %f with Magnitude: %f\n" %
              (class_re.get_freq(), class_re.get_s_par_magnitude()))

    cnt = 0

    with open(Z_PATH) as file:
        for line in file:
            # line is valid if .strip("\t") returns something with len = 3
            tmp = line.strip().split("\t")
            if len(tmp) == 3:
                # check if frequencies differs
                if abs(results[cnt].get_freq() - float(tmp[0])) < 0.0000001:
                    # place the real z value
                    results[cnt].set_z_re(float(tmp[1]))
                    # place the imaginary z value
                    results[cnt].set_z_img(float(tmp[2]))
                    # calculate magnitude of z value
                    results[cnt].calc_z_magnitude()
                    print("Freq: %f, Z real: %f and Z imaginary: %f" %
                          (results[cnt].get_freq(), results[cnt].get_z_re(), results[cnt].get_z_img()))
                cnt += 1
        file.close()

    cnt = 0

    with open(VSWR_PATH) as file:
        for line in file:
            # line is valid if .strip("\t") returns something with len = 2
            tmp = line.strip().split("\t")
            if len(tmp) == 2:
                if abs(results[cnt].get_freq() - float(tmp[0])) < 0.0000001:
                    # place VSVR
                    results[cnt].set_vsvr(float(tmp[1]))
                    print("Freq: %f and VSVR: %f" % (results[cnt].get_freq(), results[cnt].get_vsvr()))
                cnt += 1
        file.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    cnt = 2

    ws.cell(row=1, column=1, value="Frequency [GHz]")
    ws.cell(row=1, column=2, value="S_1,1 Real")
    ws.cell(row=1, column=3, value="S_1,1 Imaginary")
    ws.cell(row=1, column=4, value="S_1,1 dB")
    ws.cell(row=1, column=5, value="Z Real Ohm")
    ws.cell(row=1, column=6, value="Z Imaginary Ohm")
    ws.cell(row=1, column=7, value="Z Magnitude dB")
    ws.cell(row=1, column=8, value="VSVR")

    for el in results:
        ws.cell(row=cnt, column=1, value=el.get_freq())
        ws.cell(row=cnt, column=2, value=el.get_s_par_re())
        ws.cell(row=cnt, column=3, value=el.get_s_par_img())
        ws.cell(row=cnt, column=4, value=el.get_s_par_magnitude())
        ws.cell(row=cnt, column=5, value=el.get_z_re())
        ws.cell(row=cnt, column=6, value=el.get_z_img())
        ws.cell(row=cnt, column=7, value=el.get_z_magnitude())
        ws.cell(row=cnt, column=8, value=el.get_vsvr())
        cnt += 1

    wb.save(OUT_PATH)


if __name__ == "__main__":
    main()
