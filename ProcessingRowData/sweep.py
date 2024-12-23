import openpyxl
import math




prev = 0
results = []


class SweepData:
    def __init__(self) -> None:
        self._frequency = 0
        self._s_par_slot_9_2 = 0 # ok
        self._s_par_slot_9_0 = 0 # ok
        self._s_par_slot_8_8 = 0 # ok
        self._s_par_slot_8_6 = 0 # ok
        self._s_par_slot_8_4 = 0 # ok
        self._s_par_slot_8_2 = 0
        self._s_par_slot_8_0 = 0 # ok
        self._s_par_slot_7_8 = 0 #
        self._s_par_slot_7_6 = 0 # ok

    def set_freq(self, f: float) -> None:
        self._frequency = f

    def get_freq(self) -> float:
        return self._frequency

    def set_s_par_slot_9_2(self, val: float) -> None:
        self._s_par_slot_9_2 = val

    def set_s_par_slot_9_0(self, val: float) -> None:
        self._s_par_slot_9_0 = val

    def set_s_par_slot_8_8(self, val: float) -> None:
        self._s_par_slot_8_8 = val

    def set_s_par_slot_8_6(self, val: float) -> None:
        self._s_par_slot_8_6 = val

    def set_s_par_slot_8_4(self, val: float) -> None:
        self._s_par_slot_8_4 = val

    def set_s_par_slot_8_2(self, val: float) -> None:
        self._s_par_slot_8_2 = val

    def set_s_par_slot_8_0(self, val: float) -> None:
        self._s_par_slot_8_0 = val

    def set_s_par_slot_7_8(self, val: float) -> None:
        self._s_par_slot_7_8 = val

    def set_s_par_slot_7_6(self, val: float) -> None:
        self._s_par_slot_7_6 = val

    def get_s_par_slot_9_2(self) -> float:
        return self._s_par_slot_9_2

    def get_s_par_slot_9_0(self) -> float:
        return self._s_par_slot_9_0

    def get_s_par_slot_8_8(self) -> float:
        return self._s_par_slot_8_8

    def get_s_par_slot_8_6(self) -> float:
        return self._s_par_slot_8_6

    def get_s_par_slot_8_4(self) -> float:
        return self._s_par_slot_8_4

    def get_s_par_slot_8_2(self) -> float:
        return self._s_par_slot_8_2

    def get_s_par_slot_8_0(self) -> float:
        return self._s_par_slot_8_0

    def get_s_par_slot_7_8(self) -> float:
        return self._s_par_slot_7_8

    def get_s_par_slot_7_6(self) -> float:
        return self._s_par_slot_7_6





def main():
    ind = 0
    cnt = 0
    with open("InputTextFiles/Sweep/S parameter sweep.txt", 'r') as file:
        for line in file:
            tmp = line.strip().split("\t")
            if len(tmp) == 5:
                sweep = SweepData()
                freq = float(tmp[0])
                re = float(tmp[1])
                img = float(tmp[2])
                db = 20 * math.log10(math.sqrt(re**2+img**2))
                if ind == 0:
                    sweep.set_freq(freq)
                    sweep.set_s_par_slot_9_2(db)
                    cnt += 1
                    results.append(sweep)
                if ind == 1:
                    results[cnt].set_s_par_slot_9_0(db)
                    cnt += 1
                if ind == 2:
                    results[cnt].set_s_par_slot_8_8(db)
                    cnt += 1
                if ind == 3:
                    results[cnt].set_s_par_slot_8_4(db)
                    cnt += 1
                if ind == 4:
                    results[cnt].set_s_par_slot_8_0(db)
                    cnt += 1
                if ind == 5:
                    results[cnt].set_s_par_slot_7_6(db)
                    cnt += 1
                if ind == 6:
                    results[cnt].set_s_par_slot_8_6(db)
                    cnt += 1
                if ind == 7:
                    results[cnt].set_s_par_slot_8_2(db)
                    cnt += 1
                if ind == 8:
                    results[cnt].set_s_par_slot_7_8(db)
                    cnt += 1
                if cnt >= 1001:
                    ind += 1
                    cnt = 0
        file.close()
        for el in results:
            print("Freq: %f, slot=9.2 %f, slot=9.0 %f" % (el.get_freq(), el.get_s_par_slot_7_8(), el.get_s_par_slot_9_0()))

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="Frequency [GHz]")
        ws.cell(row=1, column=2, value="S_11 [dB] L=9.2mm")
        ws.cell(row=1, column=3, value="S_11 [dB] L=9.0mm")
        ws.cell(row=1, column=4, value="S_11 [dB] L=8.8mm")
        ws.cell(row=1, column=5, value="S_11 [dB] L=8.6mm")
        ws.cell(row=1, column=6, value="S_11 [dB] L=8.4mm")
        ws.cell(row=1, column=7, value="S_11 [dB] L=8.2mm")
        ws.cell(row=1, column=8, value="S_11 [dB] L=8.0mm")
        ws.cell(row=1, column=9, value="S_11 [dB] L=7.8mm")
        ws.cell(row=1, column=10, value="S_11 [dB] L=7.6mm")

        tmp = 2
        for el in results:
            ws.cell(row=tmp, column=1, value=el.get_freq())
            ws.cell(row=tmp, column=2, value=el.get_s_par_slot_9_2())
            ws.cell(row=tmp, column=3, value=el.get_s_par_slot_9_0())
            ws.cell(row=tmp, column=4, value=el.get_s_par_slot_8_8())
            ws.cell(row=tmp, column=5, value=el.get_s_par_slot_8_6())
            ws.cell(row=tmp, column=6, value=el.get_s_par_slot_8_4())
            ws.cell(row=tmp, column=7, value=el.get_s_par_slot_8_2())
            ws.cell(row=tmp, column=8, value=el.get_s_par_slot_8_0())
            ws.cell(row=tmp, column=9, value=el.get_s_par_slot_7_8())
            ws.cell(row=tmp, column=10, value=el.get_s_par_slot_7_6())
            tmp += 1

        wb.save("Sweep.xlsx")
if __name__ == "__main__":
   main()